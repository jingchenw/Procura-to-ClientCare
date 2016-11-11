import csv
import time
import sys
import shutil
import openpyxl
import os
from operator import itemgetter

# Prompt to type in file directory
print 'Welcome!\n'
print 'Please confirm the raw data files have been put in the "input_files" folder.(Press Enter to continue)\n'
sys.stdin.readline()

# Open and read ServiceTypeDesc
csv_file = open('res/ServiceTypeDesc.csv')
csv_reader = csv.reader(csv_file)
data = list(csv_reader)
row_count = len(data)
service_type_desc = [['' for x in range(2)] for y in range(row_count)]
csv_file = open('res/ServiceTypeDesc.csv')
csv_reader = csv.reader(csv_file)
row_number = 0
for row in csv_reader:
    service_type_desc[row_number][0] = row[0]
    service_type_desc[row_number][1] = row[1]
    row_number += 1

# Open file for row counting
print "Opening source file..."
file_name = raw_input('Please type in the raw data file name, without the file extension: ')
file_dir = 'input_files/'+str(file_name)+'.csv'
csv_file = open(file_dir)
csv_reader = csv.reader(csv_file)

# Count the INVOICE row
print "Counting the rows..."
inv_count = 0
for row in csv_reader:
    if row[0] == 'INVOICE':
        inv_count += 1

# Create 2-D List
print "Creating index..."
inv_data = [['' for x in range(27)] for y in range(inv_count)]

# Open file for data mapping
print "Mapping data..."
csv_file = open(file_dir)
csv_reader = csv.reader(csv_file)

# Map all INVOICE data to list
row_number = 0
for row in csv_reader:
    if row[0] == 'INVOICE':
        for column_count in range(27):
            inv_data[row_number][column_count] = row[column_count]
        row_number += 1

# Change date format from YYYYMMDD to DD/MM/YYYY
print "Changing date format..."
for row_number in range(inv_count):
    from_date = inv_data[row_number][1]
    conv_date = time.strptime(from_date, "%Y%m%d")
    target_date = time.strftime("%d/%m/%Y", conv_date)
    inv_data[row_number][1] = target_date

# Transfer data into BillingSum
print "Transferring data..."
billing_sum = [['' for x in range(7)] for y in range(inv_count)]
row_number = 0
for row_number in range(inv_count):
    billing_sum[row_number][0] = int(inv_data[row_number][8])      # URN
    billing_sum[row_number][1] = inv_data[row_number][20]     # COST CENTRE
    if inv_data[row_number][22] == 'MOBILITYEXP' and inv_data[row_number][21] == 'PRIVATE':
        billing_sum[row_number][2] = 'RESCONP'                # MASTER ACCT
    elif inv_data[row_number][22] == 'MOBILITYEXP':
        billing_sum[row_number][2] = 'RESCON'
    else: billing_sum[row_number][2] = inv_data[row_number][22]
    billing_sum[row_number][3] = inv_data[row_number][1]      # RECORD DATE
    billing_sum[row_number][4] = float(inv_data[row_number][12])     # AMOUNT
    billing_sum[row_number][5] = int(inv_data[row_number][11])     # INVOICE NUMBER
    if inv_data[row_number][15] == 'CC':                      # NOTES
        flag_count = 0
        for row in service_type_desc:
            if inv_data[row_number][16] == row[0]:
                if inv_data[row_number][16] == 'CC':
                    billing_sum[row_number][6] = row[1] + " " + inv_data[row_number][18] + " day/s"
                else: billing_sum[row_number][6] = row[1]
            else: flag_count += 1
        if flag_count == len(service_type_desc):
            billing_sum[row_number][6] = "Please update the Billing Service Type definition lookup table"
    elif inv_data[row_number][16] == 'HCPADJCR':
        for row in service_type_desc:
            if row[0] == inv_data[row_number][16]:
                billing_sum[row_number][6] = row[1]
    elif inv_data[row_number][16] == 'HCPADJDB':
        for row in service_type_desc:
            if row[0] == inv_data[row_number][16]:
                billing_sum[row_number][6] = row[1]
    else: billing_sum[row_number][6] = inv_data[row_number][4] + " Service Fee"

# Sum-up (Excel Macro Function)
# STEP 1 - Create a dictionary
# STEP 2 - Use URN + COST CENTRE + MASTER ACCT + RECORD DATE + INVOICE NUMBER + NOTES as key, and put AMOUNT as value
# STEP 3 - If the key has not appeared in the dictionary, add the key in, as well as the entire line of data
#          If the key has already existed in the dictionary, update the value as a sum of existing value and the current AMOUNT
# STEP 4 - Repeat STEP 2 & 3, and a final output contains the sum up of client billing amount would be generated
# STEP 5 - Transfer the final data from the dictionary into a list for further data processing
sum_up_dict = dict()
for row in billing_sum:
    key = str(row[0]) + str(row[1]) + str(row[2]) + str(row[3]) + str(row[5]) + str(row[6])
    if key in sum_up_dict.keys():
        sum_up_dict[key][4] = float("{0:.2f}".format(float(sum_up_dict[key][4]) + float(row[4])))
    else: sum_up_dict[key] = row

output_list = sum_up_dict.values()
output_list.sort(key=itemgetter(0))

print output_list

# Copy the template file into Resi Admin Folder
print "Copying template..."
shutil.copy2("output_files/CareSys-ResWorkfile-v206-Procura_3Nov2016.xlsm", "\\\Pacnsw\shared\PACData\FILE STORE\SHARED FILES\RESACC\Community Care Programs\PROCURA Billing files\CareSys_Temp.xlsm")

# Open file from Resi Admin folder
print "Final processing..."
wb = openpyxl.load_workbook("\\\Pacnsw\shared\PACData\FILE STORE\SHARED FILES\RESACC\Community Care Programs\PROCURA Billing files\CareSys_Temp.xlsm",keep_vba=True)

# Write data into the template
ws = wb.get_sheet_by_name('Paste Here')

for j in range(len(output_list)):
    ws['A' + str(j + 2)].value = output_list[j][0]
    ws['B' + str(j + 2)].value = output_list[j][1]
    ws['C' + str(j + 2)].value = output_list[j][2]
    ws['D' + str(j + 2)].value = output_list[j][3]
    ws['E' + str(j + 2)].value = output_list[j][4]
    ws['F' + str(j + 2)].value = output_list[j][5]
    ws['G' + str(j + 2)].value = output_list[j][6]

output_name = raw_input('Enter the output file name, without file extension: ') + '.xlsm'

# Save final output
print "Renaming file..."
wb.save("\\\Pacnsw\shared\PACData\FILE STORE\SHARED FILES\RESACC\Community Care Programs\PROCURA Billing files\\" + output_name)
os.remove("\\\Pacnsw\shared\PACData\FILE STORE\SHARED FILES\RESACC\Community Care Programs\PROCURA Billing files\CareSys_Temp.xlsm")